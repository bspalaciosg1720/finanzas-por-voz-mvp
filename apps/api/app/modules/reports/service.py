import csv
import io
from datetime import UTC, date, datetime, timedelta
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.modules.categories.models import Category
from app.modules.reports.schemas import (
    ReportCategory,
    ReportPeriod,
    ReportPoint,
    ReportSummary,
)
from app.modules.transactions.accounting import CONSUMPTION_ROLES, EARNED_INCOME_ROLES
from app.modules.transactions.models import Transaction
from app.modules.users.models import User


def period_bounds(anchor: date, period: ReportPeriod) -> tuple[date, date]:
    if period == "daily":
        return anchor, anchor
    if period == "weekly":
        start = anchor - timedelta(days=anchor.weekday())
        return start, start + timedelta(days=6)
    if period == "monthly":
        start = anchor.replace(day=1)
        next_month = (
            date(anchor.year + 1, 1, 1)
            if anchor.month == 12
            else date(anchor.year, anchor.month + 1, 1)
        )
        return start, next_month - timedelta(days=1)
    return date(anchor.year, 1, 1), date(anchor.year, 12, 31)


def _utc_bounds(start_date: date, end_date: date, timezone: str) -> tuple[datetime, datetime]:
    zone = ZoneInfo(timezone)
    start = datetime.combine(start_date, datetime.min.time(), zone).astimezone(UTC)
    end = datetime.combine(end_date + timedelta(days=1), datetime.min.time(), zone).astimezone(UTC)
    return start, end


def _rows(
    db: Session, user: User, start_date: date, end_date: date
) -> list[tuple[Transaction, str | None]]:
    start, end = _utc_bounds(start_date, end_date, user.timezone)
    return list(
        db.execute(
            select(Transaction, Category.name)
            .outerjoin(Category, Category.id == Transaction.category_id)
            .where(
                Transaction.user_id == user.id,
                Transaction.currency == user.default_currency,
                Transaction.status == "confirmed",
                Transaction.deleted_at.is_(None),
                Transaction.occurred_at >= start,
                Transaction.occurred_at < end,
            )
        ).all()
    )


def _series(
    rows: list[tuple[Transaction, str | None]],
    period: ReportPeriod,
    timezone: str,
) -> list[ReportPoint]:
    zone = ZoneInfo(timezone)
    grouped: dict[str, list[int]] = {}
    for item, _ in rows:
        if item.type == "income" and item.financial_role not in EARNED_INCOME_ROLES:
            continue
        if item.type == "expense" and item.financial_role not in CONSUMPTION_ROLES:
            continue
        occurred_at = item.occurred_at
        if occurred_at.tzinfo is None:
            occurred_at = occurred_at.replace(tzinfo=UTC)
        local = occurred_at.astimezone(zone)
        label = (
            local.strftime("%H:00")
            if period == "daily"
            else local.strftime("%Y-%m")
            if period == "annual"
            else local.strftime("%Y-%m-%d")
        )
        values = grouped.setdefault(label, [0, 0])
        values[0 if item.type == "income" else 1] += item.amount_minor
    return [
        ReportPoint(label=label, income_minor=values[0], expense_minor=values[1])
        for label, values in sorted(grouped.items())
    ]


def get_report(db: Session, user: User, *, period: ReportPeriod, anchor: date) -> ReportSummary:
    start_date, end_date = period_bounds(anchor, period)
    rows = _rows(db, user, start_date, end_date)
    days = (end_date - start_date).days + 1
    previous_end = start_date - timedelta(days=1)
    previous_start = previous_end - timedelta(days=days - 1)
    previous_rows = _rows(db, user, previous_start, previous_end)
    income = sum(
        item.amount_minor
        for item, _ in rows
        if item.type == "income" and item.financial_role in EARNED_INCOME_ROLES
    )
    expense = sum(
        item.amount_minor
        for item, _ in rows
        if item.type == "expense" and item.financial_role in CONSUMPTION_ROLES
    )
    previous_income = sum(
        item.amount_minor
        for item, _ in previous_rows
        if item.type == "income" and item.financial_role in EARNED_INCOME_ROLES
    )
    previous_expense = sum(
        item.amount_minor
        for item, _ in previous_rows
        if item.type == "expense" and item.financial_role in CONSUMPTION_ROLES
    )
    grouped: dict[tuple[str | None, str], int] = {}
    for item, category_name in rows:
        if item.type != "expense" or item.financial_role not in CONSUMPTION_ROLES:
            continue
        key = (
            str(item.category_id) if item.category_id else None,
            category_name or "Sin categoría",
        )
        grouped[key] = grouped.get(key, 0) + item.amount_minor
    categories = [
        ReportCategory(
            category_id=category_id,
            name=name,
            amount_minor=amount,
            percentage=round(amount / expense * 100, 1) if expense else 0,
        )
        for (category_id, name), amount in sorted(
            grouped.items(), key=lambda entry: entry[1], reverse=True
        )
    ]
    return ReportSummary(
        period=period,
        start_date=start_date,
        end_date=end_date,
        currency=user.default_currency,
        income_minor=income,
        expense_minor=expense,
        balance_minor=income - expense,
        transaction_count=len(rows),
        previous_income_minor=previous_income,
        previous_expense_minor=previous_expense,
        expense_change_percent=(
            round((expense - previous_expense) / previous_expense * 100, 1)
            if previous_expense
            else None
        ),
        categories=categories,
        series=_series(rows, period, user.timezone),
    )


def export_csv(db: Session, user: User, *, period: ReportPeriod, anchor: date) -> tuple[str, bytes]:
    start_date, end_date = period_bounds(anchor, period)
    rows = _rows(db, user, start_date, end_date)
    zone = ZoneInfo(user.timezone)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        ["fecha", "tipo", "rol_financiero", "monto", "moneda", "categoria", "descripcion", "origen"]
    )
    for item, category_name in sorted(rows, key=lambda row: row[0].occurred_at):
        occurred_at = item.occurred_at
        if occurred_at.tzinfo is None:
            occurred_at = occurred_at.replace(tzinfo=UTC)
        writer.writerow(
            [
                occurred_at.astimezone(zone).isoformat(),
                item.type,
                item.financial_role,
                item.amount_minor,
                item.currency,
                category_name or "Sin categoría",
                item.description,
                item.source,
            ]
        )
    filename = f"movimientos-{period}-{start_date}-{end_date}.csv"
    return filename, ("\ufeff" + output.getvalue()).encode("utf-8")


def _export_rows(
    db: Session, user: User, period: ReportPeriod, anchor: date
) -> tuple[date, date, list[list[str | int]]]:
    start_date, end_date = period_bounds(anchor, period)
    rows = _rows(db, user, start_date, end_date)
    zone = ZoneInfo(user.timezone)
    exported: list[list[str | int]] = []
    for item, category_name in sorted(rows, key=lambda row: row[0].occurred_at):
        occurred_at = item.occurred_at
        if occurred_at.tzinfo is None:
            occurred_at = occurred_at.replace(tzinfo=UTC)
        exported.append(
            [
                occurred_at.astimezone(zone).isoformat(),
                item.type,
                item.amount_minor,
                item.currency,
                category_name or "Sin categoría",
                item.description,
                item.source,
            ]
        )
    return start_date, end_date, exported


def export_xlsx(
    db: Session, user: User, *, period: ReportPeriod, anchor: date
) -> tuple[str, bytes]:
    start_date, end_date, rows = _export_rows(db, user, period, anchor)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimientos"
    sheet.append(["Fecha", "Tipo", "Monto", "Moneda", "Categoría", "Descripción", "Origen"])
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="245B62")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 28, "B": 12, "C": 16, "D": 10, "E": 22, "F": 38, "G": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for cell in sheet["C"][1:]:
        cell.number_format = "#,##0"
    output = io.BytesIO()
    workbook.save(output)
    return f"movimientos-{period}-{start_date}-{end_date}.xlsx", output.getvalue()


def export_pdf(db: Session, user: User, *, period: ReportPeriod, anchor: date) -> tuple[str, bytes]:
    start_date, end_date, rows = _export_rows(db, user, period, anchor)
    summary = get_report(db, user, period=period, anchor=anchor)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Reporte financiero",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Reporte financiero", styles["Title"]),
        Paragraph(f"Periodo: {start_date} — {end_date}", styles["Normal"]),
        Paragraph(
            f"Ingresos: {summary.income_minor:,} {summary.currency} · "
            f"Gastos: {summary.expense_minor:,} {summary.currency} · "
            f"Balance: {summary.balance_minor:,} {summary.currency}",
            styles["Normal"],
        ),
        Spacer(1, 6 * mm),
    ]
    table_rows: list[list[object]] = [
        ["Fecha", "Tipo", "Monto", "Moneda", "Categoría", "Descripción", "Origen"]
    ]
    for row in rows:
        table_rows.append(
            [
                str(row[0])[:16],
                row[1],
                f"{int(row[2]):,}",
                row[3],
                Paragraph(escape(str(row[4])), styles["BodyText"]),
                Paragraph(escape(str(row[5])), styles["BodyText"]),
                row[6],
            ]
        )
    table = Table(
        table_rows,
        repeatRows=1,
        colWidths=[34 * mm, 19 * mm, 27 * mm, 18 * mm, 34 * mm, 75 * mm, 24 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#245B62")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DFE4E0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F4F5F2")],
                ),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return f"reporte-{period}-{start_date}-{end_date}.pdf", output.getvalue()
